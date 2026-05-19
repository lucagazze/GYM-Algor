"""
Importa datos del Excel progreso_gym_2026 a Supabase.
Hoja LOG: columnas EJERCICIO / PR KG TOTAL / REPS / 1RM EST / FECHA PR
"""
import sys
import os
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system("pip install openpyxl -q")
    import openpyxl

try:
    from supabase import create_client, Client
except ImportError:
    print("Instalando supabase...")
    os.system("pip install supabase -q")
    from supabase import create_client, Client

# ── CONFIG ────────────────────────────────────────────────────
SUPABASE_URL = "https://omcdlfdgtnmolntmhits.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9tY2RsZmRndG5tb2xudG1oaXRzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Njc3ODA4NTksImV4cCI6MjA4MzM1Njg1OX0.rzMrnKD0Q6jwGblx2OLQcW51g8nyp5geJJqWl44FlvY"

EXCEL_PATH = r"C:\Users\lucag\Downloads\progreso_gym_2026 (1) (1).xlsx"

# Mapeo nombre Excel → nombre en DB (deben existir en exercises)
EXERCISE_MAP = {
    "Press Militar Parado (barra)": "Press Militar Parado (barra)",
    "Pecho Plano": "Pecho Plano",
    "Dips (fondos)": "Dips (fondos)",
    "Dominadas Prona": "Dominadas Prona",
    "Dominada Supina": "Dominada Supina",
    "Sentadilla": "Sentadilla",
    "Front Lever Hold (seg)": "Front Lever Hold",
    "L-Sit Hold (seg)": "L-Sit Hold",
    "Muscle Up": "Muscle Up",
}

DEFAULT_BODY_WEIGHT = 86.0

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def parse_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None

def main():
    print("Conectando a Supabase...")
    sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Cargar ejercicios de la DB
    res = sb.table("exercises").select("id,name").execute()
    exercises_db = {ex["name"]: ex["id"] for ex in (res.data or [])}
    print(f"Ejercicios en DB: {list(exercises_db.keys())}")

    # Abrir Excel
    print(f"Leyendo Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if "LOG" not in wb.sheetnames:
        print("ERROR: No se encontró la hoja LOG")
        sys.exit(1)

    ws = wb["LOG"]
    rows = list(ws.iter_rows(values_only=True))

    # Detectar fila de encabezados buscando "EJERCICIO"
    header_row = None
    for i, row in enumerate(rows):
        for cell in row:
            if str(cell).strip().upper() == "EJERCICIO":
                header_row = i
                break
        if header_row is not None:
            break

    if header_row is None:
        print("ERROR: No se encontró columna EJERCICIO")
        sys.exit(1)

    headers = [str(c).strip().upper() if c else "" for c in rows[header_row]]
    print(f"Encabezados fila {header_row + 1}: {headers}")

    # Índices de columnas
    col_cat = next((i for i, h in enumerate(headers) if h in ("", ) and i == 0), 0)
    col_ex = next((i for i, h in enumerate(headers) if "EJERCICIO" in h), None)
    col_weight = next((i for i, h in enumerate(headers) if "PR KG" in h or "KG TOTAL" in h), None)
    col_reps = next((i for i, h in enumerate(headers) if h == "REPS"), None)
    col_1rm = next((i for i, h in enumerate(headers) if "1RM" in h), None)
    col_date = next((i for i, h in enumerate(headers) if "FECHA" in h), None)

    print(f"Cols → EJ:{col_ex} PESO:{col_weight} REPS:{col_reps} 1RM:{col_1rm} FECHA:{col_date}")

    # Iterar datos desde la fila siguiente al header
    logs_to_insert = []
    skipped = 0

    for row in rows[header_row + 1:]:
        if not any(row):
            continue

        ex_name_raw = str(row[col_ex]).strip() if col_ex is not None and row[col_ex] else ""
        if not ex_name_raw or ex_name_raw.upper() in ("EJERCICIO", "NONE", ""):
            continue

        # Buscar mapeo
        ex_name_mapped = None
        for raw_key, db_name in EXERCISE_MAP.items():
            if raw_key.lower() in ex_name_raw.lower() or ex_name_raw.lower() in raw_key.lower():
                ex_name_mapped = db_name
                break

        if not ex_name_mapped:
            print(f"  SKIP sin mapeo: '{ex_name_raw}'")
            skipped += 1
            continue

        ex_id = exercises_db.get(ex_name_mapped)
        if not ex_id:
            print(f"  SKIP no en DB: '{ex_name_mapped}'")
            skipped += 1
            continue

        weight = parse_float(row[col_weight]) if col_weight is not None else None
        reps = int(parse_float(row[col_reps]) or 0) if col_reps is not None else 0
        est_1rm = parse_float(row[col_1rm]) if col_1rm is not None else None
        date_str = parse_date(row[col_date]) if col_date is not None else None

        if not date_str:
            skipped += 1
            continue

        # Calcular added_weight (total - body_weight)
        added_w = max(0.0, (weight or 0) - DEFAULT_BODY_WEIGHT) if weight else 0.0

        log = {
            "date": date_str,
            "exercise_id": ex_id,
            "added_weight": added_w,
            "body_weight": DEFAULT_BODY_WEIGHT,
            "reps": reps,
            "duration_sec": 0,
            "estimated_1rm": est_1rm or 0,
            "rir": 0,
            "is_pr": False,
            "notes": "Importado desde Excel 2026",
        }
        logs_to_insert.append(log)

    print(f"\nRegistros a importar: {len(logs_to_insert)} | Saltados: {skipped}")

    if not logs_to_insert:
        print("Nada para importar.")
        return

    confirm = input(f"\n¿Importar {len(logs_to_insert)} registros? (s/n): ").strip().lower()
    if confirm != "s":
        print("Cancelado.")
        return

    # Insertar en batches de 50
    batch_size = 50
    inserted = 0
    errors = 0
    for i in range(0, len(logs_to_insert), batch_size):
        batch = logs_to_insert[i:i + batch_size]
        try:
            res = sb.table("workout_logs").insert(batch).execute()
            inserted += len(batch)
            print(f"  Batch {i//batch_size + 1}: {len(batch)} registros ✓")
        except Exception as e:
            print(f"  Batch {i//batch_size + 1} ERROR: {e}")
            errors += len(batch)

    print(f"\n✓ Importación completa: {inserted} insertados, {errors} errores")

if __name__ == "__main__":
    main()
